import requests
import json
import pandas as pd
from pandas import ExcelFile
from pandas import ExcelWriter
import openpyxl
from os import path


# from dataclasses import dataclass


"""
script creado por julioolivares90@gmail.com
script para obtener un archivo excel con los datos de la pagina https://www.tiendacercasv.com/
"""
END_POINT = 'https://1fzqk3npw4.execute-api.us-east-1.amazonaws.com/nearby_store_stage/sv'  # endpoint  de la api de tiendacerca.com funciona con metodo post
# api de google maps para places
API_KEY = 'AIzaSyCswu8im_YgIcNBGFmRr-gRVBLqBHwXVxk'
lugar = 'Santiago Texacuangos, San Salvador'
NAME_FILE_EXCEL = 'DatosDeTiendaCerca.xlsx'  # nombre de archivo execel

# Metodo para buscar un lugar y obtener sus datos

# retorna un json con los datos de un lugar


def FindPlace(lugar, apikey):
    result = requests.get(
        "https://maps.googleapis.com/maps/api/place/findplacefromtext/json?input={0}&inputtype=textquery&fields=formatted_address,name,geometry&key={1}".format(lugar, apikey))
    return json.loads(result.content)
    pass

# Metodo para solicitar los datos a la pagina de https://www.tiendacercasv.com/


def GetDataStores(enviar_coordenadas):
    # convierte las coordenadas en una string parseada a json
    js = json.dumps(enviar_coordenadas)
    # realiza la peticion al servidor de tiendacerca.com y obtiene el json con los datos de las tiendas
    response = requests.post(END_POINT, js)
    # print(response.text)
    # convierte la respuesta en un json
    dt = json.loads(response.content)
    return dt
    pass

# lee el archivo json que contiene todos los departamentos y sus municipios


def GetDepartamentos():
    with open('Departamentos.json', 'r', encoding='utf8') as j:
        data = json.load(j)
        # print(data)
        return data
        pass  # fin with
    pass  # fin metodo
# print(dt)

# inicia todo el programa
# parametros departamentos = lista de departamentos , numero departamento = numero de departamento del 0 al 13 pensando en arraglos de programacion


def inicio_programa(departamentos, numero_departamento):
    departamentos_array = departamentos['departamentos'][numero_departamento]
    for municipio in departamentos_array['municipios']:
        # print('municipio {0} Nombre departamento {1} '.format(municipio,municipios_array['nombre']))

        dt = FindPlace('{0}, {1}'.format(
            municipio, departamentos_array['nombre']), API_KEY)
        # prepara los datos para ser enviados al servidor
        try:
            da = {
                "center": {
                    "lat": dt['candidates'][0]['geometry']['location']['lat'],
                    "lag": dt['candidates'][0]['geometry']['location']['lng']
                },
                "zoom": 10,
                "country_code": "sv",
                "east": dt['candidates'][0]['geometry']['viewport']['northeast']['lng'],
                "north": dt['candidates'][0]['geometry']['viewport']['northeast']['lat'],
                "south": dt['candidates'][0]['geometry']['viewport']['southwest']['lat'],
                "west": dt['candidates'][0]['geometry']['viewport']['southwest']['lng']
            }
            # eviar las coordenadas al server de tiendacerca y obtine los la lista de todas las tienda de un municipio
            datos_tiendas = GetDataStores(da)
            # se encarga de escribir todas las tiendas que encuentre para un municipio
            Write_data_into_file(
                departamentos_array['nombre'], municipio, datos_tiendas)
        except print('no se encontraron datos'):
            pass

        pass
    pass

# escribe en el archivo de excel los datos obtenidos


def Write_data_into_file(nombre_departamento, nombre_municipio, datos_tiendas):
    # se convertiran en los datos de las filas del excel
    """
    departamentosName = []
    municipioName = []
    ids = []
    titles = []
    address = []
    phones = []
    country_codes = []
    codes = []
    lats = []
    lngs = []
    """
   # write = ExcelWriter(NAME_FILE_EXCEL)
    # print('procesando  {0} items.....'.format(len(datos_tiendas.items())))

    for key, dt in datos_tiendas.items():
        # df = pd.read_excel(NOMBRE_EXCEL,sheet_name='Hoja 1')
        # print(dt)
        """"
        departamentosName.append(nombre_departamento)
        municipioName.append(nombre_municipio)
        ids.append(key)
        titles.append(dt['title'])
        address.append(dt['address'])
        phones.append(dt['phone'])
        country_codes.append(dt['country_code'])
        codes.append(dt['code'])
        lats.append(dt['lat'])
        lngs.append(dt['lng'])

        print('{0}---{1}'.format(key, dt['title']))
        """
        verificar_archivo(NAME_FILE_EXCEL, data=dt,
                          departamento_name=nombre_departamento, municipio_name=nombre_municipio)
        pass
    """
            df = pd.DataFrame({'departamento': departamentosName,
                                'municipio': municipioName, 'id': ids, 'title': titles, 'address': address, 'phone': phones, 'country_code': country_codes, 'code': codes, 'lat': lats, 'lng': lngs})
            df.to_excel(write,'Hoja 1',index=False)
            write.save()
            print('tarea completada')
    """
    pass
# imprime el menu de departamentos


def inprimir_nombre_departamento(departamentos):
    departamentos_array = departamentos['departamentos']
    for numero_departamento in range(len(departamentos_array)):
        print('{0} --para --{1}'.format(numero_departamento,
                                        departamentos_array[numero_departamento]['nombre']))
        pass
    pass
# verifica si el archivo ya existe y si existe lo lee


def verificar_archivo(name_file, data, departamento_name, municipio_name):
    datos = []
    if path.exists(name_file):
        wb = openpyxl.load_workbook(name_file)
       # print(data)
        datos.append(departamento_name)
        datos.append(municipio_name)
        for d in data:
            # print(data[d])
            datos.append(data[d])
            pass
        sheet = wb['Hoja1']
        # print(d[1])
        # datos.append(departamento_name, municipio_name, data['id'], data['title'], data['address'],
        #                data['phone'], data['country_code'], data['code'], data['lat'], data['lng'])
        sheet.append(datos)
        wb.save(name_file)
    pass


"""
    if path.exists(name_file):
        wb = openpyxl.load_workbook(name_file)
        sheet = wb['Hoja 1']
        for key, d in data.items():
            datos.append([departamento_name,municipio_name,key,d['title'],d['address'],d['phone'],d['country_code'],d['code'],['lat'],d['lng']])
            sheet.append(datos)
            pass
        wb.save(name_file)
        pass

    else:
        openpyxl.Workbook()
        verificar_archivo(name_file,data,departamento_name,municipio_name)
        pass
    """

if __name__ == "__main__":
    print('programa iniciado  obteniendo datos')
    dep = GetDepartamentos()
    print('----------------------------------Incio Programa Datos Tienda Cerca------------------------------------------')
    inprimir_nombre_departamento(dep)
    seguir = False

    numero_departamento = int(
        input('escribe el numero del municiipio que deseas obtener sus datos: '))

    inicio_programa(dep, numero_departamento)
    print('finalizo el programa')
    pass
